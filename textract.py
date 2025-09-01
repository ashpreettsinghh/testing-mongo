import os
import logging
import traceback
from PIL import Image
import pandas as pd
import re
import json
import uuid
from textractor import Textractor
from textractor.visualizers.entitylist import EntityList
from textractor.data.constants import TextractFeatures
import io
import inflect
from collections import OrderedDict
import boto3
import time
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
import numpy as np
from botocore.config import Config
from textractor.data.text_linearization_config import TextLinearizationConfig

# Configure logging for production
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s',
    handlers=[
        logging.FileHandler('document_processing.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Global configuration with error handling
try:
    s3 = boto3.client("s3")
    config = Config(
        read_timeout=600, 
        retries=dict(max_attempts=5)
    )
    logger.info("AWS S3 client initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize AWS S3 client: {str(e)}")
    raise

# Textractor configuration with error handling
try:
    text_config = TextLinearizationConfig(
        hide_figure_layout=False,
        title_prefix="<titles><<title>><title>",
        title_suffix="</title><</title>>",
        hide_header_layout=True,
        section_header_prefix="<headers><<header>><header>",
        section_header_suffix="</header><</header>>",
        table_prefix="<tables><table>",
        table_suffix="</table>",
        list_layout_prefix="<<list>><list>",
        list_layout_suffix="</list><</list>>",
        hide_footer_layout=True,
        hide_page_num_layout=True,
    )
    logger.info("Text linearization configuration created successfully")
except Exception as e:
    logger.error(f"Failed to create text linearization config: {str(e)}")
    raise

def safe_strip_newline(cell):
    """
    A utility function to safely strip newline characters from a cell.
    
    Parameters:
    cell: The cell value (any type)
    
    Returns:
    str: The cell value with newline characters removed, or empty string if None
    """
    try:
        if cell is None:
            return ""
        return str(cell).strip()
    except Exception as e:
        logger.warning(f"Error stripping newline from cell {cell}: {str(e)}")
        return ""

def layout_table_to_excel(document, table_id, csv_separator):    
    """
    Converts an Excel table from a document to a Pandas DataFrame, 
    handling duplicated values across merged cells with comprehensive error handling.

    Args:
        document: Document containing Excel table 
        table_id: ID of the Excel table in the document
        csv_separator: Separator for CSV string conversion

    Returns: 
        Pandas DataFrame representation of the Excel table, or None if error occurs
    """
    try:
        if not hasattr(document, 'tables') or len(document.tables) <= table_id:
            logger.error(f"Table ID {table_id} not found in document. Available tables: {len(document.tables) if hasattr(document, 'tables') else 0}")
            return None
            
        # Save the table in excel format to preserve the structure of any merged cells
        buffer = io.BytesIO()    
        document.tables[table_id].to_excel(buffer)
        buffer.seek(0)
        
        # Load workbook, get active worksheet
        wb = openpyxl.load_workbook(buffer)
        worksheet = wb.active
        
        if worksheet is None:
            logger.error(f"Failed to get active worksheet for table {table_id}")
            return None
            
        # Unmerge cells, duplicate merged values to individual cells
        all_merged_cell_ranges = list(worksheet.merged_cells.ranges)
        logger.info(f"Processing {len(all_merged_cell_ranges)} merged cell ranges for table {table_id}")
        
        for merged_cell_range in all_merged_cell_ranges:
            try:
                merged_cell = merged_cell_range.start_cell
                worksheet.unmerge_cells(range_string=merged_cell_range.coord)
                for row_index, col_index in merged_cell_range.cells:
                    cell = worksheet.cell(row=row_index, column=col_index)
                    cell.value = merged_cell.value
            except Exception as e:
                logger.warning(f"Error processing merged cell range {merged_cell_range.coord}: {str(e)}")
                continue
                
        # Convert to DataFrame with error handling
        df = pd.DataFrame(worksheet.values)
        df = df.map(safe_strip_newline)
        
        # Validate DataFrame
        if df.empty:
            logger.warning(f"Table {table_id} resulted in empty DataFrame")
            return pd.DataFrame()
            
        df_csv = df.to_csv(sep=csv_separator, index=False, header=None)
        row_count = len([x for x in df_csv.split("\n") if x.strip()])
        
        if row_count > 1:
            first_row_values = df_csv.split("\n")[0].split(csv_separator)
            if not all(value.strip() == '' for value in first_row_values): 
                row_count = 1
        
        # Attach table column names with error handling
        column_row = 0 if row_count == 1 else 1
        try:
            df.columns = df.iloc[column_row] 
            df = df[column_row+1:]
        except IndexError as e:
            logger.warning(f"Error setting column names for table {table_id}: {str(e)}. Using default columns.")
            # Use default column names if setting fails
            df.columns = [f"Column_{i}" for i in range(len(df.columns))]
            
        logger.info(f"Successfully processed table {table_id} with shape {df.shape}")
        return df
        
    except Exception as e:
        logger.error(f"Error processing table {table_id}: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None

def safe_split_list_items(items):
    """
    Safely splits the given string into a list of items, handling nested lists.

    Parameters:
    items (str): The input string containing items and possibly nested lists.

    Returns:
    list: A list containing the items extracted from the input string, empty list if error
    """
    try:
        if not isinstance(items, str):
            logger.warning(f"Expected string input, got {type(items)}. Converting to string.")
            items = str(items)
            
        parts = re.split("(<<list>><list>|</list><</list>>)", items)  
        output = []
        inside_list = False
        list_item = ""

        for p in parts:
            try:
                if p == "<<list>><list>":
                    inside_list = True    
                    list_item = p
                elif p == "</list><</list>>":
                    inside_list = False
                    list_item += p
                    output.append(list_item)
                    list_item = "" 
                elif inside_list:
                    list_item += p.strip()
                else:
                    output.extend(p.split('\n'))
            except Exception as e:
                logger.warning(f"Error processing list part '{p}': {str(e)}")
                continue
                
        logger.info(f"Successfully split list into {len(output)} items")
        return output
        
    except Exception as e:
        logger.error(f"Error in split_list_items: {str(e)}")
        return []

def safe_sub_header_content_split(string):   
    """
    Safely splits the input string by XML tags and returns a list containing the segments of text,
    excluding segments containing specific XML tags such as "<header>", "<list>", or "<table>".

    Parameters:
    string (str): The input string to be processed.

    Returns:
    list: A list containing the segments of text extracted from the input string, empty list if error
    """ 
    try:
        if not isinstance(string, str):
            logger.warning(f"Expected string input, got {type(string)}. Converting to string.")
            string = str(string)
            
        pattern = re.compile(r'<<[^>]+>>')
        segments = re.split(pattern, string)
        result = []
        
        for segment in segments:
            try:
                if segment.strip():
                    if "<header>" not in segment and "<list>" not in segment and "<table>" not in segment:
                        segment_lines = [x.strip() for x in segment.split('\n') if x.strip()]
                        result.extend(segment_lines)
                    else:
                        result.append(segment)
            except Exception as e:
                logger.warning(f"Error processing segment: {str(e)}")
                continue
                
        return result
        
    except Exception as e:
        logger.error(f"Error in sub_header_content_split: {str(e)}")
        return []

def process_document_content(document, config):
    """
    Process document content with comprehensive error handling.
    
    Args:
        document: The document object
        config: Text linearization configuration
        
    Returns:
        tuple: (document_holder, table_page) or (None, None) if critical error
    """
    try:
        if not hasattr(document, 'pages'):
            logger.error("Document object missing 'pages' attribute")
            return None, None
            
        csv_separator = "|"
        document_holder = {}
        table_page = {}
        count = 0
        unmerge_span_cells = True
        
        logger.info(f"Processing document with {len(document.pages)} pages")
        
        # Loop through each page in the document
        for page_id, page in enumerate(document.pages):
            try:
                page_text = page.get_text(config=config)
                table_count = len([word for word in page_text.split() if "<tables><table>" in word])
                
                # Validate table count
                actual_table_count = len(page.tables) if hasattr(page, 'tables') else 0
                if table_count != actual_table_count:
                    logger.warning(f"Page {page_id}: Expected {table_count} tables, found {actual_table_count}")
                
                content = page_text.split("<tables>")
                document_holder[page_id] = []    
                
                for idx, item in enumerate(content):
                    try:
                        if "<table>" in item:           
                            # Process table with error handling
                            df = None
                            if unmerge_span_cells:
                                df = layout_table_to_excel(document, count, csv_separator)
                            else:
                                try:
                                    df_csv = document.tables[count].to_pandas(use_columns=False).to_csv(
                                        header=False, index=None, sep=csv_separator
                                    )
                                    row_count = len([x for x in df_csv.split("\n") if x])
                                    
                                    if row_count > 1:
                                        first_row_values = df_csv.split("\n")[0].split(csv_separator)
                                        if not all(value.strip() == '' for value in first_row_values): 
                                            row_count = 1
                                            
                                    df = pd.read_csv(
                                        io.StringIO(df_csv), 
                                        sep=csv_separator, 
                                        header=0 if row_count == 1 else 1, 
                                        keep_default_na=False
                                    )
                                    df.rename(columns=lambda x: '' if str(x).startswith('Unnamed:') else x, inplace=True)
                                except Exception as e:
                                    logger.error(f"Error processing table {count} on page {page_id}: {str(e)}")
                                    count += 1
                                    continue
                            
                            if df is not None and not df.empty:
                                table = df.to_csv(index=None, sep=csv_separator)
                                
                                # Store table in table_page
                                if page_id in table_page:
                                    table_page[page_id].append(table)
                                else:
                                    table_page[page_id] = [table]
                                
                                # Extract table data and remaining content
                                pattern = re.compile(r'<table>(.*?)(</table>)', re.DOTALL) 
                                table_match = re.search(pattern, item)
                                
                                if table_match:
                                    remaining_content = item[table_match.end():]
                                else:
                                    remaining_content = item
                                    logger.warning(f"No table match found in item on page {page_id}")
                                
                                content[idx] = f"<<table>><table>{table}</table><</table>>"
                                count += 1
                                
                                # Check for list items in remaining content
                                if "<<list>>" in remaining_content:
                                    output = safe_split_list_items(remaining_content)
                                    output = [x.strip() for x in output if x.strip()]
                                    document_holder[page_id].extend([content[idx]] + output)           
                                else:
                                    remaining_lines = [x.strip() for x in remaining_content.split('\n') if x.strip()]
                                    document_holder[page_id].extend([content[idx]] + remaining_lines)
                            else:
                                logger.warning(f"Failed to process table {count} on page {page_id}, skipping")
                                count += 1
                        else:   
                            # Check for list items and tables in remaining content
                            if "<<list>>" in item and "<table>" not in item:   
                                output = safe_split_list_items(item)
                                output = [x.strip() for x in output if x.strip()]
                                document_holder[page_id].extend(output)
                            else:
                                item_lines = [x.strip() for x in item.split("\n") if x.strip()]
                                document_holder[page_id].extend(item_lines)
                    except Exception as e:
                        logger.error(f"Error processing content item {idx} on page {page_id}: {str(e)}")
                        continue
                        
            except Exception as e:
                logger.error(f"Error processing page {page_id}: {str(e)}")
                continue
        
        logger.info(f"Document processing completed. Pages processed: {len(document_holder)}")
        return document_holder, table_page
        
    except Exception as e:
        logger.error(f"Critical error in document processing: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None, None

def safe_document_chunking(document_holder, max_words=200):
    """
    Safely chunk document content with comprehensive error handling.
    
    Args:
        document_holder: Dictionary containing document content by page
        max_words: Maximum words per chunk
        
    Returns:
        tuple: (chunks, table_header_dict, chunk_header_mapping, list_header_dict) or None if error
    """
    try:
        if not document_holder:
            logger.error("Document holder is empty or None")
            return None, None, None, None
            
        # Flatten the nested list document_holder into a single list and Join by "\n"
        flattened_list = []
        for page_id, sublist in document_holder.items():
            try:
                if isinstance(sublist, list):
                    flattened_list.extend(sublist)
                else:
                    logger.warning(f"Page {page_id} content is not a list: {type(sublist)}")
                    flattened_list.append(str(sublist))
            except Exception as e:
                logger.warning(f"Error flattening page {page_id}: {str(e)}")
                continue
        
        if not flattened_list:
            logger.error("No content found after flattening document")
            return None, None, None, None
            
        result = "\n".join(flattened_list)
        header_split = result.split("<titles>")
        
        chunks = {}
        table_header_dict = {} 
        chunk_header_mapping = {}
        list_header_dict = {}
        
        logger.info(f"Processing {len(header_split)} title sections for chunking")
        
        # Iterate through each title section
        for title_id, items in enumerate(header_split):
            try:
                title_chunks = []
                current_chunk = []
                num_words = 0   
                table_header_dict[title_id] = {}
                chunk_header_mapping[title_id] = {}
                list_header_dict[title_id] = {}
                chunk_counter = 0
                
                for item_id, item in enumerate(items.split('<headers>')):
                    try:
                        lines = safe_sub_header_content_split(item)             
                        SECTION_HEADER = None 
                        TITLES = None
                        num_words = 0  
                        first_header_portion = True
                        
                        for line_id, line in enumerate(lines):
                            try:
                                if not line.strip():
                                    continue
                                    
                                # Extract titles
                                if "<title>" in line:   
                                    title_matches = re.findall(r'<title>(.*?)</title>', line)
                                    if title_matches:
                                        TITLES = title_matches[0].strip()
                                        line = TITLES 
                                        if re.sub(r'<[^>]+>', '', "".join(lines)).strip() == TITLES:
                                            chunk_header_mapping[title_id][chunk_counter] = lines
                                            chunk_counter += 1
                                
                                # Extract section headers
                                if "<header>" in line:   
                                    header_matches = re.findall(r'<header>(.*?)</header>', line)
                                    if header_matches:
                                        SECTION_HEADER = header_matches[0].strip()
                                        line = SECTION_HEADER    
                                        first_header_portion = True
                                
                                next_num_words = num_words + len(re.findall(r'\w+', line))  

                                # Handle regular text chunking
                                if "<table>" not in line and "<list>" not in line:
                                    current_chunk_text = "".join(current_chunk).strip()
                                    if (next_num_words > max_words and 
                                        current_chunk_text != (SECTION_HEADER or "") and 
                                        current_chunk and 
                                        current_chunk_text != (TITLES or "")):
                                        
                                        if SECTION_HEADER:
                                            if first_header_portion:
                                                first_header_portion = False                                            
                                            else:
                                                current_chunk.insert(0, SECTION_HEADER.strip())                       
                                        
                                        title_chunks.append(current_chunk[:])  # Create copy
                                        chunk_header_mapping[title_id][chunk_counter] = lines[:]  # Create copy
                                        current_chunk = []
                                        num_words = 0 
                                        chunk_counter += 1
                             
                                    current_chunk.append(line)    
                                    num_words += len(re.findall(r'\w+', line))

                                # Handle table processing
                                if "<table>" in line:
                                    try:
                                        # Get table header
                                        line_index = lines.index(line)
                                        header = ""
                                        if (line_index != 0 and 
                                            "<table>" not in lines[line_index-1] and 
                                            "<list>" not in lines[line_index-1]):
                                            header = lines[line_index-1].replace("<header>","").replace("</header>","")
                                        
                                        # Extract table content
                                        table_parts = line.split("<table>")
                                        if len(table_parts) > 1:
                                            table_content = table_parts[-1].split("</table>")[0]
                                        else:
                                            logger.warning(f"Malformed table tag in line: {line[:100]}...")
                                            continue
                                            
                                        # Process table as DataFrame
                                        try:
                                            df = pd.read_csv(io.StringIO(table_content), sep=csv_separator, 
                                                           keep_default_na=False, header=None)
                                            if df.empty:
                                                logger.warning(f"Empty table found at line {line_id}")
                                                continue
                                                
                                            df.columns = df.iloc[0]
                                            df = df[1:]
                                            df.rename(columns=lambda x: '' if str(x).startswith('Unnamed:') else x, inplace=True)
                                        except Exception as e:
                                            logger.error(f"Error creating DataFrame from table: {str(e)}")
                                            continue
                                        
                                        # Process table chunks
                                        table_chunks = []
                                        curr_chunk = [df.columns.to_list()]
                                        words = len(re.findall(r'\w+', str(current_chunk) + " " + str(curr_chunk)))
                                        
                                        # Iterate through table rows
                                        for row in df.itertuples(index=False):
                                            try:
                                                curr_chunk.append(row)         
                                                words += len(re.findall(r'\w+', str(row)))
                                                
                                                if words > max_words:                        
                                                    # Store table metadata
                                                    if chunk_counter in table_header_dict[title_id]:
                                                        table_header_dict[title_id][chunk_counter].extend([header, table_content])
                                                    else:
                                                        table_header_dict[title_id][chunk_counter] = [header, table_content]
                                                    
                                                    # Create table chunk CSV
                                                    tab_chunk = "\n".join([
                                                        csv_separator.join(str(x) for x in curr_chunk[0])
                                                    ] + [
                                                        csv_separator.join(str(x) for x in r) for r in curr_chunk[1:]
                                                    ])
                                                    
                                                    # Add header and section header if available
                                                    if header:
                                                        if (current_chunk and 
                                                            current_chunk[-1].strip().lower() == header.strip().lower()):
                                                            current_chunk.pop(-1)
                                                        
                                                        if (SECTION_HEADER and 
                                                            SECTION_HEADER.lower().strip() != header.lower().strip()):
                                                            if first_header_portion:
                                                                first_header_portion = False
                                                            else:
                                                                current_chunk.insert(0, SECTION_HEADER.strip())
                                                        
                                                        header_formatted = (header.strip() + ':' 
                                                                          if not header.strip().endswith(':') 
                                                                          else header.strip())
                                                        current_chunk.extend([header_formatted, tab_chunk])
                                                        title_chunks.append(current_chunk[:])
                                                    else:
                                                        if SECTION_HEADER:
                                                            if first_header_portion:
                                                                first_header_portion = False
                                                            else:
                                                                current_chunk.insert(0, SECTION_HEADER.strip())
                                                        current_chunk.extend([tab_chunk])
                                                        title_chunks.append(current_chunk[:])
                                                    
                                                    chunk_header_mapping[title_id][chunk_counter] = lines[:]
                                                    chunk_counter += 1
                                                    num_words = 0
                                                    current_chunk = []
                                                    curr_chunk = [curr_chunk[0]]  # Reset with headers
                                                    words = len(re.findall(r'\w+', str(curr_chunk[0])))
                                                    
                                            except Exception as e:
                                                logger.error(f"Error processing table row: {str(e)}")
                                                continue
                                        
                                        # Handle remaining table chunk
                                        if (curr_chunk != [df.columns.to_list()] and 
                                            lines.index(line) == len(lines) - 1):
                                            # Process final table chunk
                                            self._process_final_table_chunk(
                                                curr_chunk, header, table_content, SECTION_HEADER,
                                                first_header_portion, current_chunk, title_chunks,
                                                chunk_header_mapping, title_id, chunk_counter,
                                                table_header_dict, csv_separator
                                            )
                                            chunk_counter += 1
                                            num_words = 0
                                            current_chunk = []
                                            
                                    except Exception as e:
                                        logger.error(f"Error processing table on page {page_id}: {str(e)}")
                                        continue

                                # Handle list processing
                                if "<list>" in line:
                                    try:
                                        # Get list header
                                        line_index = lines.index(line)
                                        header = ""
                                        if (line_index != 0 and 
                                            "<table>" not in lines[line_index-1] and 
                                            "<list>" not in lines[line_index-1]):
                                            header = lines[line_index-1].replace("<header>","").replace("</header>","")
                                        
                                        # Extract list content
                                        list_pattern = re.compile(r'<list>(.*?)(?:</list>|$)', re.DOTALL)   
                                        list_match = re.search(list_pattern, line)
                                        
                                        if list_match:
                                            list_content = list_match.group(1)
                                            list_lines = list_content.split("\n")
                                            
                                            # Process list chunks
                                            curr_chunk = []  
                                            words = len(re.findall(r'\w+', str(current_chunk)))
                                            
                                            for list_item in list_lines:
                                                try:
                                                    if not list_item.strip():
                                                        continue
                                                        
                                                    curr_chunk.append(list_item)         
                                                    words += len(re.findall(r'\w+', list_item))
                                                    
                                                    if words >= max_words:
                                                        # Store list metadata
                                                        if chunk_counter in list_header_dict[title_id]:
                                                            list_header_dict[title_id][chunk_counter].extend([header, list_content])
                                                        else:
                                                            list_header_dict[title_id][chunk_counter] = [header, list_content]
                                                        
                                                        words = 0     
                                                        list_chunk = "\n".join(curr_chunk)
                                                        
                                                        # Add headers and process chunk
                                                        if header:
                                                            if (current_chunk and 
                                                                current_chunk[-1].strip().lower() == header.strip().lower()):
                                                                current_chunk.pop(-1)
                                                            
                                                            if (SECTION_HEADER and 
                                                                SECTION_HEADER.lower().strip() != header.lower().strip()):
                                                                if first_header_portion:
                                                                    first_header_portion = False
                                                                else:
                                                                    current_chunk.insert(0, SECTION_HEADER.strip())
                                                            
                                                            header_formatted = (header.strip() + ':' 
                                                                              if not header.strip().endswith(':') 
                                                                              else header.strip())
                                                            current_chunk.extend([header_formatted, list_chunk])
                                                            title_chunks.append(current_chunk[:])
                                                        else:
                                                            if SECTION_HEADER:
                                                                if first_header_portion:
                                                                    first_header_portion = False
                                                                else:
                                                                    current_chunk.insert(0, SECTION_HEADER.strip())
                                                            current_chunk.extend([list_chunk])
                                                            title_chunks.append(current_chunk[:])
                                                        
                                                        chunk_header_mapping[title_id][chunk_counter] = lines[:]
                                                        chunk_counter += 1
                                                        num_words = 0
                                                        current_chunk = []
                                                        curr_chunk = []
                                                        
                                                except Exception as e:
                                                    logger.error(f"Error processing list item: {str(e)}")
                                                    continue
                                            
                                            # Handle remaining list content
                                            if curr_chunk and lines.index(line) == len(lines) - 1:
                                                self._process_final_list_chunk(
                                                    curr_chunk, header, list_content, SECTION_HEADER,
                                                    first_header_portion, current_chunk, title_chunks,
                                                    chunk_header_mapping, title_id, chunk_counter,
                                                    list_header_dict
                                                )
                                                chunk_counter += 1
                                                num_words = 0
                                                current_chunk = []
                                        else:
                                            logger.warning(f"No list match found in line: {line[:100]}...")
                                            
                                    except Exception as e:
                                        logger.error(f"Error processing list on page {page_id}: {str(e)}")
                                        continue

                            except Exception as e:
                                logger.warning(f"Error processing line {line_id} in item {item_id}: {str(e)}")
                                continue
                        
                        # Handle remaining content in current chunk
                        current_chunk_text = "".join(current_chunk).strip()
                        if (current_chunk and 
                            current_chunk_text != (SECTION_HEADER or "") and 
                            current_chunk_text != (TITLES or "")):
                            
                            if SECTION_HEADER:
                                if first_header_portion:
                                    first_header_portion = False
                                else:
                                    current_chunk.insert(0, SECTION_HEADER.strip())
                            
                            title_chunks.append(current_chunk[:])
                            chunk_header_mapping[title_i
